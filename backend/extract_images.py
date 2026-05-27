from openpyxl import load_workbook
import os

EXCEL_PATH = "data/JCK 2026 Price List.xlsx"
OUTPUT_DIR = "photos"

os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = load_workbook(EXCEL_PATH)
ws = wb.active

images = ws._images

print(f"Found {len(images)} embedded images")

# Find all valid product rows
product_rows = {}

for row in range(2, ws.max_row + 1):

    unique_code = ws[f"A{row}"].value

    if unique_code:
        product_rows[row] = unique_code

sorted_rows = sorted(product_rows.keys())

# Function to find nearest product row ABOVE image
def find_closest_product_row(image_row):

    closest = None

    for row in sorted_rows:

        if row <= image_row:
            closest = row
        else:
            break

    return closest

# Extract images
for idx, image in enumerate(images):

    try:

        image_row = image.anchor._from.row + 1

        matched_row = find_closest_product_row(image_row)

        if not matched_row:
            print(f"No matching row for image {idx}")
            continue

        unique_code = product_rows[matched_row]

        image_bytes = image._data()

        ext = image.path.split(".")[-1].lower()

        if ext not in ["png", "jpg", "jpeg"]:
            ext = "png"

        save_path = os.path.join(
            OUTPUT_DIR,
            f"{unique_code}.{ext}"
        )

        with open(save_path, "wb") as f:
            f.write(image_bytes)

        print(f"Saved: {save_path}")

    except Exception as e:
        print(f"Error image {idx}: {e}")